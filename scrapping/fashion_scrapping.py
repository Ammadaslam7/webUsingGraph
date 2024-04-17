from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import openpyxl

# Initialize Chrome webdriver
webdriver_path = r'C:\Program Files (x86)\chromedriver.exe'
s = Service(webdriver_path)
driver = webdriver.Chrome(service=s)

# Open the webpage
url = 'https://finance.yahoo.com/news/asml-orders-plunge-chipmakers-pause-100726739.html/'
driver.get(url)

# Find the article content
article_content_element = driver.find_element(By.CLASS_NAME, 'caas-xray-pills-container')

# Extract the text content of the article
article_content = article_content_element.text

# Close the browser
driver.quit()

# Write the article content to an Excel file
output_file = 'fashion.xlsx'

# Load existing workbook if it exists, otherwise create a new one
try:
    wb = openpyxl.load_workbook(output_file)
except FileNotFoundError:
    wb = openpyxl.Workbook()

ws = wb.active

# Find the next available row in column A
next_row = ws.max_row + 1

# Write the article content to the next available row in column A
ws.cell(row=next_row, column=1).value = article_content

# Save the workbook
wb.save(output_file)
